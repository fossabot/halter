# src/core/services/export_to_xlnx.py
import ipaddress
import re
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from halter.core.models.area import Area
from halter.core.models.project import Network, Project

HEADERS = [
    "Канал связи",
    "Наименование оборудования",
    "IP Адрес",
    "Маска",
    "Шлюз",
    "Оборудование",
    "Сетевое имя",
    "Примечание",
]

# Взято из твоего шаблона
COL_WIDTHS = {
    "A": 11.66,
    "B": 45.33,
    "C": 14.44,
    "D": 18.55,
    "E": 7.0,
    "F": 17.0,
    "G": 20.11,
    "H": 12.55,
}
ZEBRA_FILL = PatternFill("solid", fgColor="FF92D050")
MAIN_CHANNEL = PatternFill("solid", fgColor="7EBDC2")
BACKUP_CHANNEL = PatternFill("solid", fgColor="F3DFA2")
TITLE_COLOR = PatternFill(bgColor="565656", fill_type="solid")
BORDER_MED = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
BORDER_LR = Border(left=Side(style="medium"), right=Side(style="medium"))
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
WHITE_FONT = Font(color="FFFFFF")


@dataclass
class DeviceInfo:
    name: str | None = None
    model: str | None = None
    hostname: str | None = None
    gateway: str | None = None
    note: str | None = None


def _extract_suffix_in_parentheses(text: str) -> str:
    """
    Извлекает текст из круглых скобок в конце строки и возвращает его в нижнем регистре,
    убирая лишние пробелы и артикли/предлоги, если нужно.

    Пример:
        "Сеть синхронизации ПЛК МПСА МНС НПС№2 (Резервный канал)" → "Резервный канал"
    """
    # Находим текст в последних круглых скобках
    match = re.search(r"\(([^)]+?)\)\s*$", text.strip())
    if not match:
        return ""

    extracted = match.group(1).strip()
    return extracted.capitalize()


def _ip_class_ru_letter(net: ipaddress.IPv4Network) -> str:
    first = int(str(net.network_address).split(".")[0])
    if 1 <= first <= 126:
        return "A"
    if first == 127:
        return "loopback"
    if 128 <= first <= 191:
        return "B"
    if 192 <= first <= 223:
        return "C"
    if 224 <= first <= 239:
        return "D"  # These addresses are reserved for multicasting.
    if 240 <= first <= 255:
        return "E"  # These addresses are reserved for experimental or research purposes
    return "?"


def _set_col_widths(ws: Worksheet) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _write_title_and_header(ws: Worksheet, sheet_title: str) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Сети {sheet_title}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].fill = PatternFill(
        start_color="565656", end_color="565656", fill_type="solid"
    )

    ws.append(HEADERS)
    for i, _ in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=i)
        c.font = HEADER_FONT
        c.alignment = c.alignment = CENTER if i == 1 else LEFT
        c.border = BORDER_MED
    ws.freeze_panes = "A3"


def _devices_by_ip(project: Project) -> dict[str, DeviceInfo]:
    """
    Достаём все IP устройств -> краткая инфа об устройстве.
    Ожидаем, что у устройств есть interfaces с полями ip_address, gateway и ссылкой на сеть (по имени или id).
    Этот код максимально устойчив к разной форме записи IP (с/без префикса).
    """
    mapping: dict[str, DeviceInfo] = {}
    for dev in getattr(project, "devices", []) or []:
        dname = getattr(dev, "description", None)
        dmodel = getattr(dev, "model", None)
        dhost = getattr(dev, "name", None)
        for iface in getattr(dev, "interfaces", []) or []:
            ip_raw = getattr(iface, "address", None)
            if not ip_raw:
                continue
            ip = str(ip_raw).split("/")[0].strip()
            gw = getattr(iface, "gateway", None)
            mapping[ip] = DeviceInfo(
                name=dname,
                model=dmodel,
                hostname=dhost,
                gateway=gw,
            )
    return mapping


def _networks_for_area(project: Project, area: str) -> list:
    nets: list[Network] = []
    for net in getattr(project, "networks", []):
        if area in net.area_type:
            nets.append(net)

    # Сортируем по IP-сети: преобразуем address в ip_network
    # Обработка случая, если address пустой
    def sort_key(net: Network) -> tuple:
        try:
            return (
                ipaddress.ip_network(net.address, strict=False),
                net.address,
            )
        except ValueError:
            # Если адрес некорректный или пустой — ставим в конец
            return (ipaddress.ip_network("0.0.0.0/0"), net.address)

    return sorted(nets, key=sort_key)


def _append_subnet_block(
    ws: Worksheet,
    area_name: str,
    net_obj: Network,
    dev_map: dict[str, DeviceInfo],
) -> None:
    nname = getattr(net_obj, "description", "Подсеть")
    cidr = getattr(net_obj, "address", None)
    atype = getattr(net_obj, "address_type", "IPv4 Network")
    gateway: str = getattr(net_obj, "gateway", None)
    if not cidr or atype != "IPv4 Network":
        return

    net = ipaddress.ip_network(str(cidr), strict=False)
    # Строка с заголовком подсети
    row = ws.max_row + 1
    ws.cell(
        row=row, column=1, value=_extract_suffix_in_parentheses(nname)
    ).alignment = CENTER
    ws.cell(row=row, column=1).fill = MAIN_CHANNEL
    ws.cell(row=row, column=1).alignment
    ws.cell(row=row, column=2, value=f"{nname}")
    ws.cell(row=row, column=3, value=str(net.network_address))
    ws.cell(row=row, column=4, value=f"{net.netmask}/{net.prefixlen}")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws.cell(
        row=row,
        column=5,
        value=f" Сеть класса «{_ip_class_ru_letter(net)}». {'Идентификатор VLAN: ' + str(net_obj.vlan.id) if net_obj.vlan is not None else ''}",
    )
    # рамки как в шаблоне
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = BORDER_MED

    start_merge_row = row
    row += 1

    # Зебра стартует каждый блок одинаково
    zebra_toggle = False
    for ip in net.hosts():
        ip_s = str(ip)
        info = dev_map.get(ip_s)

        ws.cell(
            row=row,
            column=2,
            value=(info.name if info and info.name else "Не используется"),
        )
        ws.cell(row=row, column=3, value=ip_s)
        ws.cell(row=row, column=4, value=f"{net.netmask}/{net.prefixlen}")
        ws.cell(row=row, column=5, value=("Шлюз" if gateway == ip_s else None))
        ws.cell(row=row, column=6, value=(info.model if info else None))
        ws.cell(row=row, column=7, value=(info.hostname if info else None))
        ws.cell(row=row, column=8, value=(info.note if info else None))

        # зебра по B..H
        if zebra_toggle:
            for c in range(2, 9):
                ws.cell(row=row, column=c).fill = ZEBRA_FILL
        zebra_toggle = not zebra_toggle

        # белый шрифт для «не используется»
        if not info:
            for c in range(2, 9):
                ws.cell(row=row, column=c).font = WHITE_FONT

        # границы (A пустая, но с LR)
        ws.cell(row=row, column=1).border = BORDER_LR
        for c in range(2, 9):
            ws.cell(row=row, column=c).border = BORDER_MED

        row += 1

    # объединяем A по блоку подсети (заголовок + все IP)
    ws.merge_cells(
        start_row=start_merge_row,
        start_column=1,
        end_row=row - 1,
        end_column=1,
    )

    # разделитель между подсетями — пустая строка, но с LR границами
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = BORDER_LR if c == 1 else BORDER_LR


def export_to_xlsx(project: Project, output_xlsx: str) -> None:
    """
    Экспорт IP-плана по твоему шаблону:
    - отдельный лист на каждый area_type из проекта;
    - в каждом листе блоки по подсетям: заголовок + все адреса по маске;
    - занятые IP заполняются данными устройства, свободные получают белый текст «не используется этот IP адрес»;
    - оформление: ширины колонок, зебра, границы, объединения ячеек как в образце.
    """
    wb = Workbook()
    # удаляем дефолтный пустой лист
    # wb.remove(wb.active)

    ws: Worksheet = wb.active  # type: ignore
    ws.title = "Network Configuration"

    # Write headers
    headers = [
        "#",
        "Network Name",
        "Address Type",
        "Address",
        "Subnet Mask",
        "VLAN",
        "Topology",
    ]
    ws.append(headers)

    # Make headers bold
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)

    i: int = 0
    # Populate data from Project instance
    for network in project.networks:
        i = i + 1
        ws.append(
            [
                i,
                network.name,  # Direct attribute access
                network.address_type,
                str(
                    ipaddress.ip_network(
                        network.address, strict=False
                    ).network_address
                )
                if network.address_type == "IPv4 Network"
                else "",
                str(
                    ipaddress.ip_network(network.address, strict=False).netmask
                )
                if network.address_type == "IPv4 Network"
                else "",
                network.vlan.id if network.vlan is not None else "",
                network.topology,
            ]
        )

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = chr(64 + col)
        ws.column_dimensions[col_letter].width = 18

    dev_map = _devices_by_ip(project)

    areas = getattr(project, "area_type", None)
    for area in areas:
        ws = wb.create_sheet(str(area))
        _set_col_widths(ws)
        _write_title_and_header(
            ws, _get_area_by_name(area, project.areas).description
        )
        nets = _networks_for_area(project, str(area))
        for net in nets:
            _append_subnet_block(ws, str(area), net, dev_map)

    wb.save(output_xlsx)


def _get_area_by_name(area_name: str, areas: list[Area]) -> Area:
    return next((a for a in areas if a.name == area_name), None)
